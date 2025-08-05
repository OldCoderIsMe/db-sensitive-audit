#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据库敏感信息审计器
Database Sensitive Information Auditor
"""

import pymysql
import pandas as pd
import os
import random
import logging
import re
import json
from typing import List, Dict, Tuple, Any, Optional
from datetime import datetime
import warnings

# 忽略pandas警告
warnings.filterwarnings('ignore')

# 配置日志
def setup_logger():
    """配置日志系统"""
    # 确保logs目录存在
    logs_dir = "logs"
    if not os.path.exists(logs_dir):
        os.makedirs(logs_dir)
    
    # 生成日志文件名（按日期）
    log_filename = os.path.join(logs_dir, f"audit_{datetime.now().strftime('%Y%m%d')}.log")
    
    # 配置日志格式
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s'
    )
    
    # 创建logger
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)
    
    # 清除已有的处理器
    if logger.handlers:
        logger.handlers.clear()
    
    # 文件处理器
    file_handler = logging.FileHandler(log_filename, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)
    
    # 控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    
    # 添加处理器
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger

logger = setup_logger()


class DatabaseAuditor:
    """数据库审计器类"""
    
    def __init__(self, output_dir: str = "audit_reports"):
        """
        初始化审计器
        
        Args:
            output_dir: 输出目录
        """
        self.output_dir = output_dir
        self.ensure_output_dir()
    
    def ensure_output_dir(self) -> None:
        """确保输出目录存在"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"创建输出目录: {self.output_dir}")
    
    def load_sensitive_rules(self) -> Dict[str, Any]:
        """
        加载敏感信息检测规则
        
        Returns:
            敏感信息检测规则配置
        """
        try:
            rules_file = os.path.join("config", "sensitive_rules.json")
            if os.path.exists(rules_file):
                with open(rules_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                logger.warning(f"敏感信息规则文件不存在: {rules_file}，使用默认规则")
                return self._get_default_rules()
        except Exception as e:
            logger.error(f"加载敏感信息规则失败: {str(e)}，使用默认规则")
            return self._get_default_rules()
    
    def _get_default_rules(self) -> Dict[str, Any]:
        """
        获取默认的敏感信息检测规则
        
        Returns:
            默认规则配置
        """
        return {
            "sensitive_rules": {
                "手机号": {
                    "field_keywords": ["phone", "mobile", "telephone"],
                    "regex_patterns": ["^1[3-9]\\d{9}$"],
                    "description": "中国手机号"
                }
            },
            "settings": {
                "enabled_rules": ["手机号"],
                "case_sensitive": False,
                "max_field_length": 100,
                "exclude_test_data": True,
                "test_patterns": ["test", "demo", "example", "sample", "fake"]
            }
        }

    def identify_sensitive_info(self, columns: List[str], record: Tuple) -> Dict[str, Any]:
        """
        识别敏感信息
        
        Args:
            columns: 字段名列表
            record: 数据记录
            
        Returns:
            包含各类敏感信息的字典
        """
        sensitive_info = {}
        
        # 加载敏感信息检测规则
        rules_config = self.load_sensitive_rules()
        rules = rules_config.get("sensitive_rules", {})
        settings = rules_config.get("settings", {})
        
        enabled_rules = settings.get("enabled_rules", [])
        case_sensitive = settings.get("case_sensitive", False)
        max_field_length = settings.get("max_field_length", 100)
        exclude_test_data = settings.get("exclude_test_data", True)
        test_patterns = settings.get("test_patterns", [])
        
        if not record or len(record) != len(columns):
            return sensitive_info
        
        for i, column in enumerate(columns):
            value = record[i]
            
            # 检查每种敏感信息类型
            for rule_name, rule_config in rules.items():
                if rule_name not in enabled_rules:
                    continue
                
                field_keywords = rule_config.get("field_keywords", [])
                regex_patterns = rule_config.get("regex_patterns", [])
                
                # 检查字段名是否匹配关键词
                is_field_match = False
                column_check = column.lower() if not case_sensitive else column
                for keyword in field_keywords:
                    keyword_check = keyword.lower() if not case_sensitive else keyword
                    if keyword_check in column_check:
                        is_field_match = True
                        break
                
                # 检查值是否匹配正则表达式（只有非None值才检查）
                is_value_match = False
                if value is not None:
                    str_value = str(value).strip()
                    
                    # 排除测试数据
                    if exclude_test_data and any(pattern.lower() in str_value.lower() for pattern in test_patterns):
                        continue
                    
                    # 限制字段长度
                    if len(str_value) > max_field_length:
                        continue
                        
                    for pattern in regex_patterns:
                        try:
                            if re.match(pattern, str_value):
                                is_value_match = True
                                break
                        except re.error as e:
                            logger.warning(f"正则表达式错误 {pattern}: {str(e)}")
                            continue
                
                # 如果字段名或值匹配，记录敏感信息
                if is_field_match or is_value_match:
                    if rule_name not in sensitive_info:
                        sensitive_info[rule_name] = {}
                    
                    # 准备显示值
                    if value is not None:
                        display_value = str(value).strip()
                        if len(display_value) > 50:
                            display_value = display_value[:50] + "..."
                    else:
                        display_value = None
                    
                    sensitive_info[rule_name][column] = {
                        "value": display_value,
                        "field_match": is_field_match,
                        "value_match": is_value_match
                    }
        
        return sensitive_info
    
    def confirm_sensitive_data(self, sensitive_info: Dict[str, Any]) -> str:
        """
        确认敏感信息是否为真实数据
        
        Args:
            sensitive_info: 敏感信息检测结果
            
        Returns:
            "是" 表示确认为真实敏感数据，"否" 表示不确认
        """
        if not sensitive_info:
            return "否"
        
        # 加载敏感信息检测规则
        rules_config = self.load_sensitive_rules()
        rules = rules_config.get("sensitive_rules", {})
        
        # 检查每种敏感信息类型
        for rule_name, detected_fields in sensitive_info.items():
            if rule_name not in rules:
                continue
                
            rule_config = rules[rule_name]
            regex_patterns = rule_config.get("regex_patterns", [])
            
            # 检查该类型下的每个字段
            for field_name, field_info in detected_fields.items():
                if not isinstance(field_info, dict):
                    continue
                    
                value = field_info.get("value")
                if value is None:
                    continue
                
                # 对值进行正则验证
                str_value = str(value).strip()
                for pattern in regex_patterns:
                    try:
                        if re.match(pattern, str_value):
                            # 找到任何一个真实匹配的值就返回"是"
                            return "是"
                    except re.error:
                        continue
        
        # 没有找到真实匹配的值
        return "否"
    
    def parse_datasource_config(self, config_text: str) -> List[Dict[str, Any]]:
        """
        解析数据源配置文本
        
        Args:
            config_text: 配置文本，格式：datasource_name,ip,port,username,password
            
        Returns:
            数据源配置列表
        """
        datasources = []
        lines = config_text.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
                
            parts = [part.strip() for part in line.split(',')]
            if len(parts) >= 5:
                datasource = {
                    'datasource_name': parts[0],
                    'ip': parts[1],
                    'port': int(parts[2]),
                    'username': parts[3],
                    'password': parts[4]
                }
                datasources.append(datasource)
                logger.info(f"解析数据源配置: {datasource['datasource_name']} - {datasource['ip']}:{datasource['port']}")
            else:
                logger.warning(f"跳过无效的配置行: {line}")
        
        return datasources
    
    def connect_database(self, datasource: Dict[str, Any]) -> Optional[pymysql.Connection]:
        """
        连接数据库
        
        Args:
            datasource: 数据源配置
            
        Returns:
            数据库连接对象或None
        """
        try:
            connection = pymysql.connect(
                host=datasource['ip'],
                port=datasource['port'],
                user=datasource['username'],
                password=datasource['password'],
                charset='utf8mb4',
                connect_timeout=10,
                read_timeout=30,
                write_timeout=30
            )
            logger.info(f"成功连接到数据库: {datasource['datasource_name']}")
            return connection
        except Exception as e:
            logger.error(f"连接数据库失败 {datasource['datasource_name']}: {str(e)}")
            return None
    
    def get_database_users(self, connection: pymysql.Connection) -> List[Dict[str, Any]]:
        """
        获取数据库用户信息
        
        Args:
            connection: 数据库连接
            
        Returns:
            用户信息列表
        """
        users = []
        try:
            with connection.cursor() as cursor:
                # 获取所有用户及其权限
                cursor.execute("""
                    SELECT 
                        User as '用户名',
                        Host as '主机',
                        SELECT_priv as '查询权限',
                        INSERT_priv as '插入权限',
                        UPDATE_priv as '更新权限',
                        DELETE_priv as '删除权限',
                        CREATE_priv as '创建权限',
                        DROP_priv as '删除权限_表',
                        RELOAD_priv as '重载权限',
                        SHUTDOWN_priv as '关闭权限',
                        PROCESS_priv as '进程权限',
                        FILE_priv as '文件权限',
                        GRANT_priv as '授权权限',
                        REFERENCES_priv as '引用权限',
                        INDEX_priv as '索引权限',
                        ALTER_priv as '修改权限',
                        SHOW_DB_priv as '显示数据库权限',
                        SUPER_priv as '超级权限',
                        CREATE_TMP_TABLE_priv as '创建临时表权限',
                        LOCK_TABLES_priv as '锁表权限',
                        EXECUTE_priv as '执行权限',
                        REPL_SLAVE_priv as '复制从权限',
                        REPL_CLIENT_priv as '复制客户端权限'
                    FROM mysql.user
                    ORDER BY User, Host
                """)
                
                columns = [desc[0] for desc in cursor.description]
                for row in cursor.fetchall():
                    user_info = dict(zip(columns, row))
                    # 转换Y/N为是/否
                    for key, value in user_info.items():
                        if value == 'Y':
                            user_info[key] = '是'
                        elif value == 'N':
                            user_info[key] = '否'
                    users.append(user_info)
                
                logger.info(f"获取到 {len(users)} 个数据库用户")
                
        except Exception as e:
            logger.error(f"获取数据库用户失败: {str(e)}")
        
        return users
    
    def get_databases(self, connection: pymysql.Connection) -> List[str]:
        """
        获取所有数据库名称
        
        Args:
            connection: 数据库连接
            
        Returns:
            数据库名称列表
        """
        databases = []
        try:
            with connection.cursor() as cursor:
                cursor.execute("SHOW DATABASES")
                for row in cursor.fetchall():
                    db_name = row[0]
                    # 跳过系统数据库
                    if db_name not in ['information_schema', 'performance_schema', 'mysql', 'sys']:
                        databases.append(db_name)
                
                logger.info(f"获取到 {len(databases)} 个业务数据库")
                
        except Exception as e:
            logger.error(f"获取数据库列表失败: {str(e)}")
        
        return databases
    
    def get_table_info(self, connection: pymysql.Connection, database: str) -> List[Dict[str, Any]]:
        """
        获取数据库中所有表的信息
        
        Args:
            connection: 数据库连接
            database: 数据库名称
            
        Returns:
            表信息列表
        """
        table_info = []
        try:
            with connection.cursor() as cursor:
                # 切换到指定数据库
                cursor.execute(f"USE `{database}`")
                
                # 获取所有表名
                cursor.execute("SHOW TABLES")
                tables = [row[0] for row in cursor.fetchall()]
                
                logger.info(f"数据库 {database} 中有 {len(tables)} 个表")
                
                for table in tables:
                    try:
                        # 获取表字段信息
                        cursor.execute(f"DESCRIBE `{table}`")
                        columns = [row[0] for row in cursor.fetchall()]
                        
                        # 获取表记录总数
                        cursor.execute(f"SELECT COUNT(*) FROM `{table}`")
                        total_count = cursor.fetchone()[0]
                        
                        # 获取随机一条记录并转换为JSON格式（如果表不为空）
                        column_value_json = "{}"
                        phone_numbers_json = "{}"
                        
                        if total_count > 0 and columns:
                            try:
                                # 使用LIMIT和OFFSET获取随机记录
                                random_offset = random.randint(0, max(0, total_count - 1))
                                cursor.execute(f"SELECT * FROM `{table}` LIMIT 1 OFFSET {random_offset}")
                                record = cursor.fetchone()
                                if record:
                                    # 创建字段名和值的字典
                                    column_value_dict = {}
                                    for i, column in enumerate(columns):
                                        if i < len(record):
                                            value = record[i]
                                            # 处理不同数据类型
                                            if value is None:
                                                column_value_dict[column] = None
                                            elif isinstance(value, (int, float, bool)):
                                                column_value_dict[column] = value
                                            else:
                                                # 转换为字符串，限制长度
                                                str_value = str(value)
                                                if len(str_value) > 100:
                                                    str_value = str_value[:100] + "..."
                                                column_value_dict[column] = str_value
                                    
                                    # 转换为JSON字符串
                                    column_value_json = json.dumps(column_value_dict, ensure_ascii=False, separators=(',', ':'))
                                    
                                    # 识别敏感信息
                                    sensitive_info = self.identify_sensitive_info(columns, record)
                                    sensitive_info_json = json.dumps(sensitive_info, ensure_ascii=False, separators=(',', ':'))
                                    
                                    # 确认敏感信息
                                    sensitive_confirmed = self.confirm_sensitive_data(sensitive_info)
                                    
                            except Exception as e:
                                logger.warning(f"获取表 {table} 随机记录失败: {str(e)}")
                                column_value_json = '{"error": "获取失败"}'
                                sensitive_info_json = '{"error": "获取失败"}'
                                sensitive_confirmed = "否"
                        elif columns:
                            # 表为空但有字段，显示字段结构
                            empty_dict = {column: None for column in columns}
                            column_value_json = json.dumps(empty_dict, ensure_ascii=False, separators=(',', ':'))
                            
                            # 对空表也检查字段名是否包含敏感信息关键词
                            empty_record = tuple([None] * len(columns))
                            sensitive_info = self.identify_sensitive_info(columns, empty_record)
                            sensitive_info_json = json.dumps(sensitive_info, ensure_ascii=False, separators=(',', ':'))
                            sensitive_confirmed = self.confirm_sensitive_data(sensitive_info)
                        
                        # 每张表创建一条记录
                        table_info.append({
                            '表名': table,
                            '字段名和值': column_value_json,
                            '敏感信息': sensitive_info_json,
                            '敏感信息确认': sensitive_confirmed,
                            '总条数': total_count
                        })
                        
                        logger.debug(f"处理表 {table}: {len(columns)} 个字段, {total_count} 条记录")
                        
                    except Exception as e:
                        logger.error(f"处理表 {table} 失败: {str(e)}")
                        # 添加错误记录
                        table_info.append({
                            '表名': table,
                            '字段名和值': f'{{"error": "{str(e)}"}}',
                            '敏感信息': '{"error": "获取失败"}',
                            '敏感信息确认': "否",
                            '总条数': 0
                        })
                
        except Exception as e:
            logger.error(f"获取数据库 {database} 表信息失败: {str(e)}")
        
        return table_info
    
    def _apply_conditional_formatting(self, worksheet, dataframe, sheet_type='database'):
        """
        应用条件格式化到Excel工作表
        
        Args:
            worksheet: openpyxl工作表对象
            dataframe: pandas数据框
            sheet_type: sheet类型，'database' 或 'users'
        """
        try:
            from openpyxl.styles import Font, PatternFill
            
            # 创建红色加粗字体样式
            red_bold_font = Font(color="FF0000", bold=True)
            
            if sheet_type == 'database':
                # 数据库sheet：只格式化"敏感信息确认"列
                sensitive_confirm_col = None
                for idx, col_name in enumerate(dataframe.columns):
                    if col_name == '敏感信息确认':
                        sensitive_confirm_col = idx + 1  # Excel列从1开始
                        break
                
                if sensitive_confirm_col is not None:
                    # 遍历数据行（跳过标题行）
                    for row_idx in range(2, len(dataframe) + 2):  # Excel行从1开始，数据从第2行开始
                        cell = worksheet.cell(row=row_idx, column=sensitive_confirm_col)
                        if cell.value == '是':
                            cell.font = red_bold_font
            
            elif sheet_type == 'users':
                # users sheet：格式化所有包含"是"值的单元格
                for row_idx in range(2, len(dataframe) + 2):  # 跳过标题行
                    for col_idx in range(1, len(dataframe.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if cell.value == '是':
                            cell.font = red_bold_font
            
            elif sheet_type == 'audit':
                # 审计结果sheet：根据风险等级着色
                from openpyxl.styles import PatternFill
                
                # 创建不同风险等级的样式
                high_risk_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 浅红色背景
                medium_risk_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # 浅黄色背景
                high_risk_font = Font(color="CC0000", bold=True)  # 深红色字体
                medium_risk_font = Font(color="FF6600", bold=True)  # 橙色字体
                
                # 找到风险等级列
                risk_level_col = None
                for idx, col_name in enumerate(dataframe.columns):
                    if col_name == '风险等级':
                        risk_level_col = idx + 1
                        break
                
                if risk_level_col is not None:
                    # 遍历数据行并应用格式
                    for row_idx in range(2, len(dataframe) + 2):
                        risk_cell = worksheet.cell(row=row_idx, column=risk_level_col)
                        
                        if risk_cell.value == '高':
                            # 高风险：整行浅红色背景，风险等级列深红色字体
                            for col_idx in range(1, len(dataframe.columns) + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = high_risk_fill
                            risk_cell.font = high_risk_font
                            
                        elif risk_cell.value == '中':
                            # 中风险：整行浅黄色背景，风险等级列橙色字体
                            for col_idx in range(1, len(dataframe.columns) + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = medium_risk_fill
                            risk_cell.font = medium_risk_font
                    
        except ImportError:
            logger.warning("openpyxl样式模块导入失败，跳过条件格式化")
        except Exception as e:
            logger.warning(f"应用条件格式化失败: {str(e)}")
    
    def _add_hyperlinks_to_audit_sheet(self, worksheet, dataframe):
        """
        为审计结果sheet的检查项列添加超链接
        
        Args:
            worksheet: openpyxl工作表对象
            dataframe: pandas数据框
        """
        try:
            from openpyxl.styles import Font
            
            # 找到检查项列
            check_item_col = None
            for idx, col_name in enumerate(dataframe.columns):
                if col_name == '检查项':
                    check_item_col = idx + 1  # Excel列从1开始
                    break
            
            if check_item_col is None:
                return
            
            # 创建超链接字体样式
            link_font = Font(color="0000FF", underline="single")
            
            # 遍历数据行并添加超链接
            for row_idx in range(2, len(dataframe) + 2):  # 跳过标题行
                cell = worksheet.cell(row=row_idx, column=check_item_col)
                check_item_value = cell.value
                
                if check_item_value and check_item_value != '用户权限':
                    # 对数据库名称添加超链接到对应sheet
                    # Excel sheet名称长度限制为31字符
                    sheet_name = check_item_value[:31] if len(check_item_value) > 31 else check_item_value
                    # 尝试使用标准的Excel内部链接格式
                    from openpyxl.worksheet.hyperlink import Hyperlink
                    cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet_name}'!A1")
                    cell.font = link_font
                elif check_item_value == '用户权限':
                    # 对用户权限添加超链接到用户权限sheet
                    from openpyxl.worksheet.hyperlink import Hyperlink
                    cell.hyperlink = Hyperlink(ref=cell.coordinate, location="'用户权限'!A1")
                    cell.font = link_font
                    
        except ImportError:
            logger.warning("openpyxl样式模块导入失败，跳过超链接添加")
        except Exception as e:
            logger.warning(f"添加超链接失败: {str(e)}")
    
    def _generate_audit_summary(self, users: List[Dict[str, Any]], 
                               databases_info: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
        """
        生成审计结果汇总
        
        Args:
            users: 用户信息列表
            databases_info: 数据库信息字典
            
        Returns:
            审计结果列表
        """
        audit_results = []
        
        # 1. 敏感信息风险汇总 - 每张表一条记录
        for database, table_info in databases_info.items():
            for table in table_info:
                if table.get('敏感信息确认') == '是':
                    # 解析敏感信息详情
                    sensitive_info = {}
                    try:
                        import json
                        sensitive_info = json.loads(table.get('敏感信息', '{}'))
                    except:
                        pass
                    
                    if sensitive_info:
                        # 汇总该表的所有敏感信息类型
                        sensitive_types = list(sensitive_info.keys())
                        sensitive_fields = []
                        sample_values = []
                        
                        for sensitive_type, fields in sensitive_info.items():
                            for field_name, details in fields.items():
                                sensitive_fields.append(f"{field_name}({sensitive_type})")
                                if details.get('value'):
                                    value = str(details.get('value', ''))
                                    sample_values.append(value[:10] + '...' if len(value) > 10 else value)
                        
                        # 生成汇总描述
                        types_str = '、'.join(sensitive_types)
                        fields_str = '、'.join(sensitive_fields[:3])  # 最多显示3个字段
                        if len(sensitive_fields) > 3:
                            fields_str += f'等{len(sensitive_fields)}个字段'
                        
                        audit_results.append({
                            '风险类型': '敏感信息',
                            '风险等级': '高',
                            '检查项': database,
                            '表名': table.get('表名', ''),
                            '字段名': fields_str,
                            '敏感类型': types_str,
                            '风险描述': f'表 {table.get("表名", "")} 包含敏感信息: {types_str}',
                            '检测值': '、'.join(sample_values[:2]) + ('...' if len(sample_values) > 2 else ''),
                            '记录总数': table.get('总条数', 0),
                            '建议': f'对包含{types_str}的字段进行加密或脱敏处理'
                        })
        
        # 2. 高危权限风险汇总
        high_risk_permissions = [
            '超级权限', '文件权限', '关闭权限', '重载权限', 
            '进程权限', '授权权限', '复制从权限', '复制客户端权限'
        ]
        
        for user in users:
            username = user.get('用户名', '')
            host = user.get('主机', '')
            
            # 检查高危权限
            user_high_risks = []
            for perm in high_risk_permissions:
                if user.get(perm) == '是':
                    user_high_risks.append(perm)
            
            if user_high_risks:
                audit_results.append({
                    '风险类型': '权限风险',
                    '风险等级': '高' if '超级权限' in user_high_risks else '中',
                    '检查项': '用户权限',
                    '表名': '-',
                    '字段名': '-',
                    '敏感类型': '数据库权限',
                    '风险描述': f'用户 {username}@{host} 拥有高危权限: {", ".join(user_high_risks)}',
                    '检测值': f'{len(user_high_risks)}个高危权限',
                    '记录总数': '-',
                    '建议': '根据最小权限原则，移除不必要的高危权限'
                })
            
            # 检查通配符主机权限
            if host == '%' and any(user.get(perm) == '是' for perm in ['查询权限', '插入权限', '更新权限', '删除权限']):
                audit_results.append({
                    '风险类型': '权限风险',
                    '风险等级': '中',
                    '检查项': '用户权限',
                    '表名': '-',
                    '字段名': '-',
                    '敏感类型': '主机权限',
                    '风险描述': f'用户 {username} 允许从任意主机(%)连接并具有数据操作权限',
                    '检测值': '通配符主机权限',
                    '记录总数': '-',
                    '建议': '限制用户只能从特定主机连接，避免使用通配符(%)'
                })
        
        # 3. 按风险等级排序
        risk_priority = {'高': 1, '中': 2, '低': 3}
        audit_results.sort(key=lambda x: (risk_priority.get(x['风险等级'], 99), x['风险类型']))
        
        return audit_results

    def generate_excel_report(self, datasource_name: str, users: List[Dict[str, Any]], 
                            databases_info: Dict[str, List[Dict[str, Any]]]) -> str:
        """
        生成Excel报告
        
        Args:
            datasource_name: 数据源名称
            users: 用户信息列表
            databases_info: 数据库信息字典
            
        Returns:
            生成的Excel文件路径
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"{datasource_name}_{timestamp}.xlsx"
        excel_path = os.path.join(self.output_dir, excel_filename)
        
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # 第一个sheet：审计结果汇总
                audit_results = self._generate_audit_summary(users, databases_info)
                if audit_results:
                    audit_df = pd.DataFrame(audit_results)
                    audit_df.to_excel(writer, sheet_name='审计结果', index=False)
                    
                    # 应用条件格式 - 审计结果sheet的风险等级着色
                    self._apply_conditional_formatting(writer.book['审计结果'], audit_df, sheet_type='audit')
                    
                    # 添加检查项列的超链接
                    self._add_hyperlinks_to_audit_sheet(writer.book['审计结果'], audit_df)
                    
                    logger.info(f"写入审计结果 sheet: {len(audit_results)} 条风险记录")
                else:
                    # 创建空的审计结果sheet
                    empty_audit_df = pd.DataFrame(columns=[
                        '风险类型', '风险等级', '检查项', '表名', '字段名', 
                        '敏感类型', '风险描述', '检测值', '记录总数', '建议'
                    ])
                    empty_audit_df.to_excel(writer, sheet_name='审计结果', index=False)
                    logger.info("创建空的审计结果 sheet")
                
                # 第二个sheet：用户权限信息
                if users:
                    users_df = pd.DataFrame(users)
                    users_df.to_excel(writer, sheet_name='用户权限', index=False)
                    
                    # 应用条件格式 - users sheet中的"是"值红色加粗
                    self._apply_conditional_formatting(writer.book['用户权限'], users_df, sheet_type='users')
                    
                    logger.info(f"写入用户信息 sheet: {len(users)} 条记录")
                else:
                    # 创建空的用户sheet
                    empty_df = pd.DataFrame(columns=['用户名', '主机', '权限信息'])
                    empty_df.to_excel(writer, sheet_name='用户权限', index=False)
                
                # 为每个数据库创建sheet
                for database, table_info in databases_info.items():
                    if table_info:
                        db_df = pd.DataFrame(table_info)
                        # Excel sheet名称长度限制为31字符
                        sheet_name = database[:31] if len(database) > 31 else database
                        db_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # 应用条件格式 - 敏感信息确认列红色加粗
                        self._apply_conditional_formatting(writer.book[sheet_name], db_df, sheet_type='database')
                        
                        logger.info(f"写入数据库 {database} sheet: {len(table_info)} 条记录")
                    else:
                        # 创建空的数据库sheet
                        empty_df = pd.DataFrame(columns=['表名', '字段名和值', '敏感信息', '敏感信息确认', '总条数'])
                        sheet_name = database[:31] if len(database) > 31 else database
                        empty_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.info(f"Excel报告生成成功: {excel_path}")
            return excel_path
            
        except Exception as e:
            logger.error(f"生成Excel报告失败: {str(e)}")
            raise
    
    def audit_datasource(self, datasource: Dict[str, Any]) -> Optional[str]:
        """
        审计单个数据源
        
        Args:
            datasource: 数据源配置
            
        Returns:
            生成的Excel文件路径或None
        """
        logger.info(f"开始审计数据源: {datasource['datasource_name']}")
        
        # 连接数据库
        connection = self.connect_database(datasource)
        if not connection:
            return None
        
        try:
            # 获取用户信息
            users = self.get_database_users(connection)
            
            # 获取所有数据库
            databases = self.get_databases(connection)
            
            # 获取每个数据库的表信息
            databases_info = {}
            for database in databases:
                logger.info(f"处理数据库: {database}")
                table_info = self.get_table_info(connection, database)
                databases_info[database] = table_info
            
            # 生成Excel报告
            excel_path = self.generate_excel_report(
                datasource['datasource_name'], 
                users, 
                databases_info
            )
            
            logger.info(f"数据源 {datasource['datasource_name']} 审计完成")
            return excel_path
            
        except Exception as e:
            logger.error(f"审计数据源 {datasource['datasource_name']} 失败: {str(e)}")
            return None
        finally:
            connection.close()
    
    def audit_multiple_datasources(self, config_text: str) -> List[str]:
        """
        审计多个数据源
        
        Args:
            config_text: 数据源配置文本
            
        Returns:
            生成的Excel文件路径列表
        """
        logger.info("开始批量审计数据源")
        
        # 解析配置
        datasources = self.parse_datasource_config(config_text)
        if not datasources:
            logger.warning("没有找到有效的数据源配置")
            return []
        
        # 审计每个数据源
        excel_files = []
        for datasource in datasources:
            excel_path = self.audit_datasource(datasource)
            if excel_path:
                excel_files.append(excel_path)
        
        logger.info(f"批量审计完成，生成了 {len(excel_files)} 个Excel报告")
        return excel_files


def main():
    """主函数 - 用于测试"""
    # 示例配置
    config_text = """
# 数据源配置示例
# datasource_name,ip,port,username,password
test_db1,localhost,3306,root,password123
test_db2,example.com,3306,admin,admin123
    """
    
    auditor = DatabaseAuditor()
    excel_files = auditor.audit_multiple_datasources(config_text)
    
    print(f"生成的Excel报告文件:")
    for file_path in excel_files:
        print(f"  - {file_path}")


if __name__ == "__main__":
    main()